import os
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()
client = OpenAI()

OUTPUT_DIR = "generated_decodable_stories_two_phase"
os.makedirs(OUTPUT_DIR, exist_ok=True)


LESSON_FRY_LIMITS = {
    35: 40,    # K mid: First 100 lists 1–4
    48: 60,    # K end: + Second 100 lists 1–2
    57: 80,    # Grade 1 beginning
    80: 120,   # Grade 1 end
    91: 240,   # Grade 2 beginning
    108: 240   # Grade 2 end
}

LESSON_GRADE = {
    35: "K",
    48 : "K",
    60: "1",
    80: "1",
    91: "2",
    120: "2"
}

LESSON_PHASE = {
    35: ("K", "mid"), 48: ("K", "end"), 60: ("1", "beginning"), 80: ("1", "end"),
    91: ("2", "beginning"), 120: ("2", "end")
}

STORY_EXPECTATIONS = {
    ("K", "mid"): {"sentences": "8–10", "target_repeats": "about 5"},
    ("K", "end"): {"sentences": "12–15", "target_repeats": "about 5"},
    ("1", "beginning"): {"sentences": "15–18", "target_repeats": "about 8"},
    ("1", "mid"): {"sentences": "18–22", "target_repeats": "about 8"},
    ("1", "end"): {"sentences": "22–25", "target_repeats": "about 8"},
    ("2", "beginning"): {"sentences": "25–28", "target_repeats": "about 10"},
    ("2", "mid"): {"sentences": "28–32", "target_repeats": "about 10"},
    ("2", "end"): {"sentences": "32–35", "target_repeats": "about 10"}
}


def load_fry_words(filepath="Word Lists/1000words.txt", limit=100):
    with open(filepath, "r", encoding="utf-8") as f:
        words = [line.strip().lower() for line in f if line.strip()]
    return words[:limit]

def load_phonics_lesson(filepath="phonics_lessons.xlsx", lesson_num=35):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[1]
    row = ws[lesson_num + 1]
    rule = row[0].value
    words_raw = row[1].value
    target_words = [w.strip().lower() for w in words_raw.split(",") if w.strip()]
    return rule, target_words

def load_previous_phonics_words(filepath="phonics_lessons.xlsx", lesson_num=35):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[1]
    review_words = set()
    for ln in range(1, lesson_num):
        row = ws[ln + 1]
        words_raw = row[1].value
        if words_raw:
            review_words.update(w.strip().lower() for w in words_raw.split(",") if w.strip())
    return list(review_words)


def generate_story_outline(fry_words, review_words, target_words, phonics_class, grade, phase, sentence_range, target_repeat_guidance):
    prompt = f"""
Plan a short decodable story aligned to UFLI phonics lesson {phonics_class}.
Use only Fry sight words (grade-appropriate) and words from previous phonics lessons.
Include target words naturally several times.

Story expectations:
- Grade {grade}, {phase} of school year
- Total story sentences: {sentence_range}
- Target words repetition: {target_repeat_guidance}

Structure:
- Beginning: introduce characters and setting
- Middle: simple problem, goal, or event
- End: problem solved or goal achieved
- Keep ONE main idea throughout
- Each sentence should connect logically to the previous

Page structure:
- Kindergarten: ~1 sentence per page
- Grade 1: ~2–3 sentences per page
- Grade 2: ~4–5 sentences per page

Examples for tone and sentence structure:
Min has a kit.
Kim has a lid.
They dig in the big pit.
“Dig, Min! Dig, Kim!”
A bug is in the pit.
The bug is big and red.
Min can lift the lid.
Kim can pin the bug in the kit.
The bug will not slip.
“Look, Kim! We did it!” said Min.
Kim had a big grin.
They put the bug in the kit.
Min, Kim, and the bug sit in the sun.
They had fun and did a big job.

Jan had a cat.
The cat was Sam.
Sam sat on Jan’s lap.
Jan had a map.
“It is a map of the park,” said Jan.
“I can run and hop at the park!”
Sam ran to the map.
Sam sat on it.
“Oh, Sam!” said Jan. “That is my map!”
Jan got the map.
Sam ran to the mat.
Jan ran to Sam.
“Let’s go, Sam!” said Jan.
Jan and Sam ran and ran.
They ran up a path.
They ran past a big rock.
At last, they sat and had a nap.
Jan had fun.
Sam had fun.

Output a JSON array of short plot points (one per sentence).
"""
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()


def generate_decodable_story(fry_words, review_words, target_words, outline_json, phonics_class, grade, phase, sentence_range, target_repeat_guidance):
    prompt = f"""
Write a short decodable story based on this outline: {outline_json}

Rules:
- Use ONLY words from these sources: Fry words ({fry_words}) and previous phonics words ({review_words})
- Target words ({target_words}) must appear naturally {target_repeat_guidance}
- Keep total sentences: {sentence_range}
- Follow grade {grade}, {phase} page sentence guidelines
- Each sentence should be short, simple, and decodable
- Keep ONE main idea
- Characters act consistently
- No long, abstract, or multi-syllable words
- Model tone and simplicity on these examples:

Min has a kit.
Kim has a lid.
They dig in the big pit.
“Dig, Min! Dig, Kim!”
A bug is in the pit.
The bug is big and red.
Min can lift the lid.
Kim can pin the bug in the kit.
The bug will not slip.
“Look, Kim! We did it!” said Min.
Kim had a big grin.
They put the bug in the kit.
Min, Kim, and the bug sit in the sun.
They had fun and did a big job.

Jan had a cat.
The cat was Sam.
Sam sat on Jan’s lap.
Jan had a map.
“It is a map of the park,” said Jan.
“I can run and hop at the park!”
Sam ran to the map.
Sam sat on it.
“Oh, Sam!” said Jan. “That is my map!”
Jan got the map.
Sam ran to the mat.
Jan ran to Sam.
“Let’s go, Sam!” said Jan.
Jan and Sam ran and ran.
They ran up a path.
They ran past a big rock.
At last, they sat and had a nap.
Jan had fun.
Sam had fun.

Output the story as plain text, one sentence per line.
"""
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
    )
    return response.choices[0].message.content.strip()


def main():
    lessons = [35, 48, 60, 80, 91, 120]

    for lesson_num in lessons:
        print(f"Generating story for UFLI lesson {lesson_num}...")

        # Fry words
        fry_limit = LESSON_FRY_LIMITS.get(lesson_num, 40)
        fry_words = load_fry_words(limit=fry_limit)

        # Previous phonics words
        review_words = load_previous_phonics_words("phonics_lessons.xlsx", lesson_num=lesson_num)

        # Lesson rule and target words
        rule, target_words = load_phonics_lesson("phonics_lessons.xlsx", lesson_num=lesson_num)

        # Grade and phase
        if lesson_num in LESSON_PHASE:
            grade, phase = LESSON_PHASE[lesson_num]
        else:
            grade, phase = LESSON_GRADE.get(lesson_num, "K"), "mid"

        # Story expectations
        sentence_range = STORY_EXPECTATIONS[(grade, phase)]["sentences"]
        target_repeat_guidance = STORY_EXPECTATIONS[(grade, phase)]["target_repeats"]

        # outline
        outline_json = generate_story_outline(
            fry_words, review_words, target_words, lesson_num, grade, phase, sentence_range, target_repeat_guidance
        )

        # full story
        story_text = generate_decodable_story(
            fry_words, review_words, target_words, outline_json, lesson_num, grade, phase, sentence_range, target_repeat_guidance
        )

        # Save story
        story_dir = os.path.join(OUTPUT_DIR, f"Lesson_{lesson_num}")
        os.makedirs(story_dir, exist_ok=True)
        with open(os.path.join(story_dir, "story.txt"), "w", encoding="utf-8") as f:
            f.write(f"UFLI Lesson {lesson_num}: {rule}\n\n")
            f.write(story_text)

        print(f"Saved story for lesson {lesson_num}\n")


if __name__ == "__main__":
    main()






